from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import numpy as np
import io, os

@csrf_exempt
def img_to_excel(request):
    if request.method == "POST" and request.FILES.get("image"):
        try:
            # Read the uploaded image
            image_file = request.FILES["image"]
            image = Image.open(image_file)

            # Convert image to RGB list
            rgb_list = image_to_rgb(image)

            # Convert RGB list to Excel
            output_dir = "output"
            os.makedirs(output_dir, exist_ok=True)
            output_file = os.path.join(output_dir, "rgb_colors.xlsx")
            rgb_to_excel(rgb_list, output_file)
            
            # Return output directory
            return JsonResponse({"message": "Excel file created successfully!", "file_path": output_file})
        
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    
    return JsonResponse({"error": "Invalid request method."}, status=405)

def image_to_rgb(image):
    # Ensure the image is in RGB mode
    if image.mode != "RGB":
        image = image.convert("RGB")
    
    # Convery image to a NumPy array
    rgb_array = np.array(image) # Shape: (rows, columns, 3)

    # Convert NumPy array to a Python list for JSON serialization
    rgb_list = rgb_array.tolist()

    return rgb_list

def rgb_to_excel(rgb_values, output_file="output.xlsx"):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "RGB Values"

    # Start writing to the Excel sheet
    current_row = 1 # Start from the first row

    # Iterate over the rows in the 3D list
    for row in rgb_values:
        for col_idx, (r, g, b) in enumerate(row, start=1):
            # Calculate the starting row for this RGB triplet
            base_row = current_row
            col = col_idx # Column stays constant for this RGB triplet

            # Write and style the Red row
            red_cell = ws.cell(row=base_row, column=col)
            red_cell.value = str(r)
            red_cell.fill = PatternFill(start_color=f"{r:02X}0000", end_color=f"{r:02X}0000", fill_type="solid")

            # Write and style the Green row
            green_cell = ws.cell(row=base_row + 1, column=col)
            green_cell.value = str(g)
            green_cell.fill = PatternFill(start_color=f"00{g:02X}00", end_color=f"00{g:02X}00", fill_type="solid")

            # Write and style the Blue row
            blue_cell = ws.cell(row=base_row + 2, column=col)
            blue_cell.value = str(b)
            blue_cell.fill = PatternFill(start_color=f"0000{b:02X}", end_color=f"0000{b:02X}", fill_type="solid")
        
        # Increment `current_row` to start the next row group
        current_row += 3
    
    wb.save(output_file)